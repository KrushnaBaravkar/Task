"""
make_monthly_boxplot_inplace.py
Reads 'Derived_Variables' sheet and creates:
 - 'Monthly_Boxplot_Data' sheet (wide table Jan..Dec with raw values)
 - 'Monthly_Boxplot_Figure' sheet with the saved boxplot image inserted.
It edits the same Excel file (creates a backup first).
"""

import os
import shutil
import calendar
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage

# ---------- USER SETTINGS ----------
INPUT_FILE = r"C:\Users\krush\Downloads\68_240003017_NagarjunaSagar_Telangana–AndhraPradesh_2000_to_2013 (1)(AutoRecovered).xlsx"  # <<-- set full path to your Excel file here
SOURCE_SHEET = "Derived_Variables"   # sheet name that has Month Name and % Filling
MONTH_COL_NAME = "Month Name"        # column header exactly as in Excel
VALUE_COL_NAME = "% Filling With Storage"
OUTPUT_WIDE_SHEET = "Monthly_Boxplot_Data"
OUTPUT_PLOT_SHEET = "Monthly_Boxplot_Figure"
PLOT_FILENAME = "monthly_boxplot.png"
# -----------------------------------

# 1) safety check + backup
if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

backup = INPUT_FILE + ".backup.xlsx"
shutil.copy2(INPUT_FILE, backup)
print(f"Backup created: {backup}")

# 2) read the source sheet into pandas
df = pd.read_excel(INPUT_FILE, sheet_name=SOURCE_SHEET, engine="openpyxl")

# 3) validate columns
if MONTH_COL_NAME not in df.columns:
    raise KeyError(f"Month column '{MONTH_COL_NAME}' not found in sheet '{SOURCE_SHEET}'.\nFound columns: {list(df.columns)}")
if VALUE_COL_NAME not in df.columns:
    raise KeyError(f"Value column '{VALUE_COL_NAME}' not found in sheet '{SOURCE_SHEET}'.")

# helper: normalize month to 3-letter abbreviations (Jan, Feb, ...)
def normalize_month(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s == "":
        return None
    low = s.lower()
    # direct map of full and abbr names
    for i in range(1,13):
        if low == calendar.month_name[i].lower() or low == calendar.month_abbr[i].lower():
            return calendar.month_abbr[i]
    # numeric strings like "2" or "02"
    try:
        n = int(s)
        if 1 <= n <= 12:
            return calendar.month_abbr[n]
    except Exception:
        pass
    # try parsing as date
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if not pd.isna(dt):
            return calendar.month_abbr[dt.month]
    except Exception:
        pass
    # fallback first 3 letters
    if len(low) >= 3:
        first3 = low[:3]
        for i in range(1,13):
            if first3 == calendar.month_abbr[i].lower():
                return calendar.month_abbr[i]
    return None

# 4) apply normalization and clean
df["_MONTH_ABBR"] = df[MONTH_COL_NAME].apply(normalize_month)
bad = df["_MONTH_ABBR"].isna().sum()
if bad > 0:
    print(f"Warning: {bad} rows couldn't be parsed as months and will be skipped.")

df_vals = df.loc[df["_MONTH_ABBR"].notna(), ["_MONTH_ABBR", VALUE_COL_NAME]].copy()
df_vals[VALUE_COL_NAME] = pd.to_numeric(df_vals[VALUE_COL_NAME], errors="coerce")
df_vals = df_vals.dropna(subset=[VALUE_COL_NAME])

# 5) create month -> list mapping
months = [calendar.month_abbr[i] for i in range(1,13)]
month_lists = {m: df_vals.loc[df_vals["_MONTH_ABBR"] == m, VALUE_COL_NAME].tolist() for m in months}
max_len = max((len(lst) for lst in month_lists.values()), default=0)

# 6) build wide DataFrame
wide_dict = {}
for m in months:
    lst = month_lists[m]
    if len(lst) < max_len:
        lst = lst + [np.nan] * (max_len - len(lst))
    wide_dict[m] = lst
wide_df = pd.DataFrame(wide_dict)

# 7) open workbook and remove existing target sheets if present
wb = openpyxl.load_workbook(INPUT_FILE)
if OUTPUT_WIDE_SHEET in wb.sheetnames:
    wb.remove(wb[OUTPUT_WIDE_SHEET])
if OUTPUT_PLOT_SHEET in wb.sheetnames:
    wb.remove(wb[OUTPUT_PLOT_SHEET])
wb.save(INPUT_FILE)
wb.close()

# 8) write the wide_df into the workbook as a new sheet
# use pandas ExcelWriter with openpyxl; if_sheet_exists requires pandas>=1.3 but we removed sheet above so we can just append
with pd.ExcelWriter(INPUT_FILE, engine="openpyxl", mode="a") as writer:
    wide_df.to_excel(writer, sheet_name=OUTPUT_WIDE_SHEET, index=False)

print(f"'{OUTPUT_WIDE_SHEET}' sheet written to {INPUT_FILE} (columns Jan..Dec).")

# 9) create a boxplot (matplotlib) and save image
plt.figure(figsize=(12,6))
ax = wide_df.plot.box(grid=True, showmeans=True)
ax.set_title("Monthly Distribution of Reservoir Storage (% Filling)")
ax.set_ylabel("Reservoir Storage (% Filling)")
# Show month labels exactly in order
ax.set_xticklabels(months)
plt.tight_layout()
plt.savefig(PLOT_FILENAME, dpi=200)
plt.close()
print(f"Boxplot saved as: {PLOT_FILENAME}")

# 10) insert the image into a new sheet
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.create_sheet(OUTPUT_PLOT_SHEET)
img = XLImage(PLOT_FILENAME)
# optionally: set image size, e.g., img.width = 900
ws.add_image(img, "A1")
wb.save(INPUT_FILE)
wb.close()
print(f"Inserted plot image into sheet '{OUTPUT_PLOT_SHEET}'. File saved: {INPUT_FILE}")

print("Done. Open the workbook in MS Excel and check sheets:")
print(f" - {OUTPUT_WIDE_SHEET}  (wide table to use for Box & Whisker)")
print(f" - {OUTPUT_PLOT_SHEET}  (embedded plot image)")
print("A backup of the original file was made with '.backup.xlsx' appended.")
